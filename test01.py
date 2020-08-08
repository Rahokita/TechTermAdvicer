import os
import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

excel_filename = 'example.xlsx'
file_passname = os.path.dirname(__file__)
excel_filename = file_passname + '/' + excel_filename

wb = openpyxl.load_workbook(excel_filename)

#シート名を出力する(12.3.2)
sheetnames = wb.sheetnames
print(sheetnames)

#アクティブシートを取得する
print(wb.active)

#シートからセルを取得する(12.3.3)
WS1 = wb['Sheet1']
print(WS1['B1'].value)
print(WS1.cell(row=1,column=2).value)

#シートのデータサイズ（行数、列数）を取得する
print(WS1.max_row)
print(WS1.max_column)

#列名称と番号の変換
print(get_column_letter(2))
print(column_index_from_string('A'))

#シートから複数の行と列を取得する(12.3.5)
print(tuple(WS1['A1':'C3']))

for row_of_cell_objects in WS1['A1':'C3']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate,cell_obj.value)
    print('-- End of Row --')
